import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from src.core.models import EmployeeAnnualReport
from src.core.config import GlobalConfig
from typing import List, Dict
from datetime import date

class ExcelReportGenerator:
    """
    Excel 报表生成器。
    Presentation 层组件，只负责将业务模型数据渲染为格式化的 Excel 报表。
    """
    
    def __init__(self, output_path: str):
        self.output_path = output_path
        self.wb = Workbook()
        self.ws = self.wb.active
        
        target_year = GlobalConfig.get_year()
        self.ws.title = f"{target_year}年员工年休假统计表"
        
        # 预定义样式
        self.font_bold = Font(bold=True, name='宋体', size=11)
        self.font_normal = Font(name='宋体', size=11)
        self.align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    def generate(self, reports: List[EmployeeAnnualReport], active_holidays: Dict[int, List[date]]):
        """执行报表生成逻辑。"""
        print(f"正在生成报表，包含 {len(reports)} 条员工记录...")
        self._setup_headers(active_holidays)
        self._fill_data(reports, active_holidays)
        self.wb.save(self.output_path)
        print(f"报表已保存至: {self.output_path}")

    def _setup_headers(self, active_holidays: Dict[int, List[date]]):
        """配置多级表头结构。"""
        target_year = GlobalConfig.get_year()
        
        # 第一行标题
        self.ws.merge_cells('B2:E2')
        self.ws['B2'] = f"{target_year}年员工年休假统计表"
        
        # 固定列头
        # 移除“年休假天数（期初）”列
        static_headers = ["序号", "姓名", "部门", "入职日期"]
        col_idx = 2
        for h in static_headers:
            self.ws.cell(row=4, column=col_idx, value=h)
            self.ws.merge_cells(start_row=4, start_column=col_idx, end_row=5, end_column=col_idx)
            col_idx += 1
        
        # 月度动态列头
        self.month_col_map = {}   # 记录每月的起始列索引，用于填充数据
        self.holiday_col_map = {} # 记录节假日日期列的索引

        for m in range(1, 13):
            holidays = active_holidays.get(m, [])
            
            # 判断当月是否有补假规则 (目前仅2月)
            has_bonus_rule = (m == 2)
            
            # 基础列：结转、休假
            # 如果有补假规则，则额外增加“补假”列
            base_cols_count = 3 if has_bonus_rule else 2
            
            # 宽度 = 基础列 + N列节假日
            width = base_cols_count + len(holidays)
            
            # 月份大标题（第4行）
            start_col = col_idx
            end_col = col_idx + width - 1
            c = self.ws.cell(row=4, column=start_col, value=f"{m}月")
            if width > 1:
                self.ws.merge_cells(start_row=4, start_column=start_col, end_row=4, end_column=end_col)
            
            # 子列头（第5行）：结转、休假、(补假)
            self.ws.cell(row=5, column=col_idx, value="结转")
            col_opening = col_idx
            col_idx += 1
            
            self.ws.cell(row=5, column=col_idx, value="休假")
            col_leave = col_idx
            col_idx += 1
            
            col_bonus = None
            if has_bonus_rule:
                self.ws.cell(row=5, column=col_idx, value="补假")
                col_bonus = col_idx
                col_idx += 1
            
            # 记录列映射：Month -> {opening: col, leave: col, bonus: col}
            self.month_col_map[m] = {'opening': col_opening, 'leave': col_leave, 'bonus': col_bonus}
            
            # 子列头（第5行）：具体节假日
            for h_date in holidays:
                date_str = f"{h_date.month}.{h_date.day}"
                self.ws.cell(row=5, column=col_idx, value=date_str)
                self.holiday_col_map[h_date] = col_idx
                col_idx += 1
        
        # 汇总列头
        # 更新为：合计已休、剩余未休、节假日未加班天数、补假(年度总)、备注
        summary_headers = ["合计已休", "剩余未休", "节假日未加班天数", "补假", "备注"]
        self.summary_col_map = {}
        for h in summary_headers:
            self.ws.cell(row=4, column=col_idx, value=h)
            self.ws.merge_cells(start_row=4, start_column=col_idx, end_row=5, end_column=col_idx)
            self.summary_col_map[h] = col_idx
            col_idx += 1

    def _fill_data(self, reports: List[EmployeeAnnualReport], active_holidays: Dict[int, List[date]]):
        """填充员工业务数据。"""
        current_row = 6
        for report in reports:
            emp = report.employee
            
            # 填充固定列
            self._set_cell(current_row, 2, emp.index)
            self._set_cell(current_row, 3, emp.name)
            self._set_cell(current_row, 4, emp.department)
            self._set_cell(current_row, 5, emp.join_date)
            
            # 移除期初余额填充
            
            # 填充月度数据
            for m in range(1, 13):
                record = report.monthly_records.get(m)
                cols = self.month_col_map[m]
                
                # 填充结转 (opening_balance) 和 休假 (actual_leave_days)
                opening_val = record.opening_balance if record else None
                leave_val = record.actual_leave_days if record else None
                
                # 填充补假 (bonus_days) - 仅当有补假列时
                bonus_val = None
                if cols['bonus']:
                    if record:
                        try:
                            b = record.bonus_days
                            if b > 0:
                                bonus_val = b
                        except AttributeError:
                            pass
                
                self._set_cell(current_row, cols['opening'], opening_val)
                self._set_cell(current_row, cols['leave'], leave_val)
                
                if cols['bonus']:
                    self._set_cell(current_row, cols['bonus'], bonus_val)
                
                # 填充节假日状态
                holidays = active_holidays.get(m, [])
                for h_date in holidays:
                    status = ''
                    if record:
                        status = record.holiday_statuses.get(h_date, '')
                    
                    if h_date in self.holiday_col_map:
                        self._set_cell(current_row, self.holiday_col_map[h_date], status)
            
            # 填充汇总数据
            self._set_cell(current_row, self.summary_col_map["合计已休"], report.total_leave_taken)
            self._set_cell(current_row, self.summary_col_map["剩余未休"], report.remaining_balance)
            self._set_cell(current_row, self.summary_col_map["节假日未加班天数"], report.total_holiday_leave)
            
            # total_bonus 也是 computed_field
            total_bonus_val = 0.0
            try:
                total_bonus_val = report.total_bonus
            except AttributeError:
                pass
            self._set_cell(current_row, self.summary_col_map["补假"], total_bonus_val)
            
            self._set_cell(current_row, self.summary_col_map["备注"], report.notes)
            
            current_row += 1

    def _set_cell(self, row, col, value):
        """辅助方法：设置单元格值并应用边框。"""
        c = self.ws.cell(row=row, column=col, value=value)
        c.border = self.border_thin
        c.alignment = self.align_center
